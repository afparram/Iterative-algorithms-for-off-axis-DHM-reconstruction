from functions_evaluation import functions_evaluation
from AlgoritmoDualAnnealing import AlgoritmosOptimizacion
import numpy as np
import pandas as pd
import os
import random
from multiprocessing import Pool, cpu_count
from itertools import product

if __name__ == '__main__':
    SEED = 42
    np.random.seed(SEED)
    random.seed(SEED)
    n_cores = cpu_count()

    # Ruta de archivo de los hologramas
    path =  os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(path, "..", "PruebasMuestras")
    hologramas = os.listdir(folder_path)

    filename = 'resultados_dual_annealing.xlsx'
    mode = 'a' if os.path.exists(filename) else 'w'

    if mode == 'a':
        from openpyxl import load_workbook
        book = load_workbook(filename)
    else:
        book = None

    with pd.ExcelWriter(filename, engine='openpyxl', mode=mode) as writer:
        if book:
            writer.book = book
            existing_sheets = book.sheetnames
        else:
            existing_sheets = []

        for filename_holo in hologramas:
            print(f"Procesando holograma: {filename_holo}")
            # Ruta absoluta del archivo del holograma
            holo_path = os.path.join(folder_path, filename_holo)

            # Crear objeto que permita leer el holograma, obtener los parámetros utilizados (lambda y dxy)
            # Además de obtener todas las funciones para la reconstrucción de fase.
            mi_holograma = functions_evaluation(holo_path)

            """
                Valor inicial de la reconstrucción de fase
                Sacar FT al holograma y filtrar con una máscara circular centrada en el punto máximo discreto.
                El método devuelve la información compleja del holograma filtrado y los puntos máximos discretos.
            """
            holo_filtered, fx_max, fy_max = mi_holograma.circular_filter(visual=False, scale = 1)
            mi_holograma.ang_spectrum(fx_max, fy_max, holo_filtered, visual=False)

            '''
                Algoritmos de reconstrucción de fase
            '''

            # Configuración de límites
            bounds = (np.array([fx_max - 1, fy_max - 1]), np.array([fx_max + 1, fy_max + 1]))

            # Crear instancia y ejecutar optimización
            algoOpt = AlgoritmosOptimizacion(
                mi_holograma,
                holo_filtered,
                [fx_max, fy_max],
                bounds=bounds,
                tol=1e-3,
                max_iter=7
            )

            hyperparams = {
            'initial_temp': [5230, 2500, 7500],
            'restart_temp_ratio': [2e-05, 2e-3, 2e-7],
            'visit': [2.62, 1.5, 2.0],
            'accept': [-5, -5e-1, -5e-2]
            }

            param_combinations = list(product(
                hyperparams['initial_temp'],
                hyperparams['restart_temp_ratio'],
                hyperparams['visit'],
                hyperparams['accept']
            ))

            with Pool(n_cores) as pool:
                resultados_hyper = dict(pool.map(algoOpt.evaluar_combinacion, param_combinations))

            base_name = f"{mi_holograma.n}x{mi_holograma.m}_{mi_holograma.lambda_}_{mi_holograma.dxy}"
            sheet_name = base_name
            counter = 1
            # mientras haya colisión, añadimos sufijo
            while sheet_name in existing_sheets:
                sheet_name = f"{base_name}_{counter}"
                counter += 1

            existing_sheets.append(sheet_name)

            # Guardar resultados en una una hoja de excel
            df = pd.DataFrame.from_dict(resultados_hyper, orient='index')
            df.to_excel(writer, sheet_name = sheet_name)