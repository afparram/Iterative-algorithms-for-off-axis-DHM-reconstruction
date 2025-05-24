from functions_evaluation import functions_evaluation
from AlgoritmosOptimizacion_V3 import AlgoritmosOptimizacion
import numpy as np
import pandas as pd
import os
import random

if __name__ == '__main__':
    SEED = 42
    np.random.seed(SEED)
    random.seed(SEED)

    # Ruta de archivo de los hologramas
    path =  os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(path, "..", "PruebasMuestras")
    hologramas = os.listdir(folder_path)

    filename = 'resultados_algoritmos_4.xlsx'
    mode = 'a' if os.path.exists(filename) else 'w'

    if mode == 'a':
        from openpyxl import load_workbook
        book = load_workbook(filename)
    else:
        book = None

    with pd.ExcelWriter(filename, engine='openpyxl', mode=mode) as writer:
        if book:
            writer.book = book
            existing_sheets = book.sheetnames
        else:
            existing_sheets = []

        for filename_holo in hologramas:
            print(f"Procesando holograma: {filename_holo}")
            # Ruta absoluta del archivo del holograma
            holo_path = os.path.join(folder_path, filename_holo)

            # Crear objeto que permita leer el holograma, obtener los parámetros utilizados (lambda y dxy)
            # Además de obtener todas las funciones para la reconstrucción de fase.
            mi_holograma = functions_evaluation(holo_path)

            """
                Valor inicial de la reconstrucción de fase
                Sacar FT al holograma y filtrar con una máscara circular centrada en el punto máximo discreto.
                El método devuelve la información compleja del holograma filtrado y los puntos máximos discretos.
            """
            holo_filtered, fx_max, fy_max = mi_holograma.circular_filter(visual=False, scale = 1)
            mi_holograma.ang_spectrum(fx_max, fy_max, holo_filtered, visual=False)

            '''
                Algoritmos de reconstrucción de fase
            '''

            # Configuración de límites
            bounds = (
                np.array([fx_max - 1, fy_max - 1]),
                np.array([fx_max + 1, fy_max + 1])
            )

            # Crear instancia y ejecutar optimización
            algoOpt = AlgoritmosOptimizacion(
                mi_holograma,
                holo_filtered,
                [fx_max, fy_max],
                bounds=bounds,
                tol=1e-3,
                max_iter=7
            )

            # Ejecutar todos los algoritmos
            resultados = algoOpt.ejecutar_todos()

            # Mostrar resultados
            tabla = algoOpt.mostrar_resultados(resultados)
            print('\n')

            number = filename_holo.split('_')[0]

            base_name = f"{mi_holograma.n}x{mi_holograma.m}_{number}_{mi_holograma.lambda_}_{mi_holograma.dxy}"
            sheet_name = base_name
            counter = 1
            # mientras haya colisión, añadimos sufijo
            while sheet_name in existing_sheets:
                sheet_name = f"{base_name}_{counter}"
                counter += 1

            existing_sheets.append(sheet_name)

            # Guardar resultados en una una hoja de excel
            df_tabla = pd.DataFrame(tabla, columns=["Método", "fx", "fy", "Costo", "Tiempo (s)"])
            df_tabla.to_excel(writer, sheet_name = sheet_name, index = False)